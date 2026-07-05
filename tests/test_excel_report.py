"""The spec's contract for the Excel report: the workbook opens, the
expected sheets exist, and values in the file match the marts they came
from — proven here against synthetic frames through the same pure
build_workbook the production run uses."""

import io

import pandas as pd
from openpyxl import load_workbook

from reporting.excel_report import BANDS, PIVOT_COLS, Z_COLUMNS, build_workbook, quarter_tag


def _frames():
    movers = []
    for band in BANDS:
        for i in range(4):
            movers.append({
                "peer_band": band, "cert": 1000 + i, "bank_name": f"{band} Bank {i}",
                "composite": 1.0 + i * 0.1, "composite_prior": 0.8, "delta": 0.2 + i * 0.1,
                **{c: (i - 2) * 1.5 for c in Z_COLUMNS},
            })
    return {
        "latest": "2026-03-31",
        "prior": "2025-12-31",
        "kpis": pd.DataFrame([{
            "banks_reporting": 1059, "active_banks": 1045, "combined_assets_tn": 25.3,
            "median_roa_pct": 1.22, "median_nim_pct": 3.61, "median_equity_pct": 10.4,
        }]),
        "h8": pd.DataFrame([
            {"obs_date": "2026-06-17", "series_title": "Bank Credit", "value_billions": 18500.0},
            {"obs_date": "2026-06-24", "series_title": "Bank Credit", "value_billions": 18510.0},
        ]),
        "movers": pd.DataFrame(movers),
    }


def _reload(frames):
    buf = io.BytesIO()
    build_workbook(frames).save(buf)
    buf.seek(0)
    return load_workbook(buf)


def test_workbook_opens_with_expected_sheets():
    wb = _reload(_frames())
    assert wb.sheetnames == ["Summary", "$1B-$10B", "$10B-$100B", "over $100B",
                             "PivotData", "Pivot", "Methodology"]


def test_pivot_data_is_a_real_table_matching_the_code():
    frames = _frames()
    wb = _reload(frames)
    ws = wb["PivotData"]
    table = ws.tables["MoversData"]
    # drift guard: the table's columns are what every live formula references
    assert [c.name for c in table.tableColumns] == PIVOT_COLS
    assert table.ref.endswith(str(len(frames["movers"]) + 1))
    # delta is a native formula, not a baked value
    assert ws.cell(row=2, column=PIVOT_COLS.index("delta") + 1).value == "=D2-E2"


def test_pivot_table_survives_a_round_trip():
    wb = _reload(_frames())
    pivots = wb["Pivot"]._pivots
    assert len(pivots) == 1
    assert [f.name for f in pivots[0].cache.cacheFields] == PIVOT_COLS
    assert pivots[0].cache.refreshOnLoad  # Excel rebuilds the cache on open


def test_summary_formulas_are_native_and_prefixed():
    wb = _reload(_frames())
    text = [str(c.value) for row in wb["Summary"].iter_rows() for c in row
            if isinstance(c.value, str) and c.value.startswith("=")]
    assert any("COUNTIFS(MoversData[peer_band]" in f for f in text)
    assert any("_xlfn.XLOOKUP(" in f for f in text)   # Excel-365 fns need the prefix on disk
    assert any("_xlfn.MAXIFS(" in f for f in text)
    band = wb["$1B-$10B"]
    assert str(band["E5"].value).startswith("=C5-D5"[:3])  # Change column is a formula


def test_spot_check_values_match_the_frames():
    wb = _reload(_frames())
    summary = wb["Summary"]
    # three values, per the spec: a KPI, an H.8 observation, and a mover's composite
    assert summary["B4"].value == 1059                       # banks reporting
    assert summary.cell(row=13, column=2).value == 18500.0   # first H.8 week, Bank Credit
    band = wb["$1B-$10B"]
    # risers are sorted by delta descending: Bank 3 (delta 0.5, composite 1.3) leads
    assert band["A4"].value == "Bank"
    assert band["A5"].value == "$1B-$10B Bank 3"
    assert band["C5"].value == 1.3


def test_methodology_stays_neutral():
    wb = _reload(_frames())
    text = " ".join(str(c.value) for row in wb["Methodology"].iter_rows() for c in row if c.value)
    assert "not an assessment" in text and "not a prediction" in text
    assert "will fail" not in text and "risky" not in text


def test_quarter_tag():
    assert quarter_tag("2026-03-31") == "2026Q1"
    assert quarter_tag("2025-12-31") == "2025Q4"
