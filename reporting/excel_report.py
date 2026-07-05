"""The quarterly peer report as a downloadable workbook: sector summary and
H.8 trend, one tab per size band with the quarter's biggest composite-score
movers under z-score color scales, and the methodology in the same neutral
language as the site. Generated on every deploy from the marts, uploaded to
the public reports bucket as bank_peer_report_<YYYYQn>.xlsx and latest.xlsx.

Split on purpose: fetch_frames does the BigQuery reads, build_workbook is
pure frames-in/workbook-out so the tests can prove sheet structure and
spot-check values without a warehouse.

Usage: uv run python -m reporting.excel_report
"""

from __future__ import annotations

import logging
import os
import sys
import tempfile
from pathlib import Path

import pandas as pd
from dotenv import load_dotenv
from google.cloud import bigquery, storage
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.pivot.cache import (
    CacheDefinition,
    CacheField,
    CacheSource,
    SharedItems,
    WorksheetSource,
)
from openpyxl.pivot.fields import Text
from openpyxl.pivot.table import (
    DataField,
    FieldItem,
    Location,
    PivotField,
    PivotTableStyle,
    RowColField,
    TableDefinition,
)
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

log = logging.getLogger(__name__)

BANDS = ["$1B-$10B", "$10B-$100B", ">$100B"]
MOVERS_PER_SIDE = 15

Z_COLUMNS = {
    "z_uninsured_share": "Uninsured share z",
    "z_brokered_share": "Brokered share z",
    "z_securities_share": "Securities z",
    "z_asset_growth_3y": "3y growth z",
    "z_nim_trend": "NIM trend z",
    "z_equity_ratio": "Equity z",
}

# the Excel Table behind the pivot and every live formula in the workbook;
# order defines the sheet layout, names must match the movers frame
PIVOT_COLS = ["peer_band", "bank_name", "cert", "composite", "composite_prior",
              "delta"] + list(Z_COLUMNS)

METHODOLOGY = """How to read this report

Every FDIC-insured bank above $1B in assets is compared against banks of
similar size each quarter: $1B-$10B, $10B-$100B, and over $100B. Six metrics
- uninsured-deposit share, brokered-deposit share, securities as a share of
assets, three-year asset growth, the four-quarter net interest margin trend,
and equity over assets - each become a z-score against the median and MAD of
the bank's size band that quarter, winsorized to [-5, +5]. The composite is
the unweighted mean of the available scores.

A high composite means one thing: this bank's numbers sit unusually far from
its size group on these six metrics. It is a peer-relative statistic computed
from public quarterly filings, not an assessment of any bank's condition and
not a prediction. Movers are the quarter's largest changes in composite
score, in both directions; a bank can move because its own numbers changed
or because its peer group's did.

Data: FDIC BankFind Suite API (quarterly filings) and the Federal Reserve
H.8 release (weekly aggregates). Dollar amounts in the FDIC data are
reported in thousands; this report shows billions. The full methodology,
lineage, and the 2023 backtest of the method live on the dashboard."""


def fetch_frames(client: bigquery.Client, marts: str) -> dict:
    """Everything the workbook needs, in five queries."""
    latest, prior = [r[0] for r in client.query_and_wait(
        f"SELECT DISTINCT report_date FROM `{marts}.fct_bank_quarters` ORDER BY 1 DESC LIMIT 2"
    )]

    kpis = client.query_and_wait(f"""
        SELECT
            count(*)                                        AS banks_reporting,
            countif(d.is_active)                            AS active_banks,
            round(sum(f.total_assets) / 1e9, 1)             AS combined_assets_tn,
            round((SELECT DISTINCT percentile_cont(roa_pct, 0.5) OVER ()
                   FROM `{marts}.fct_bank_quarters` WHERE report_date = '{latest}'), 2)
                                                            AS median_roa_pct,
            round((SELECT DISTINCT percentile_cont(net_interest_margin_pct, 0.5) OVER ()
                   FROM `{marts}.fct_bank_quarters` WHERE report_date = '{latest}'), 2)
                                                            AS median_nim_pct,
            round((SELECT DISTINCT percentile_cont(equity_to_assets, 0.5) OVER ()
                   FROM `{marts}.fct_bank_quarters` WHERE report_date = '{latest}') * 100, 1)
                                                            AS median_equity_pct
        FROM `{marts}.fct_bank_quarters` f
        JOIN `{marts}.dim_banks` d USING (cert)
        WHERE f.report_date = '{latest}'
        """).to_dataframe()

    h8 = client.query_and_wait(f"""
        SELECT obs_date, series_title, round(value_billions, 1) AS value_billions
        FROM `{marts}.stg_fred__h8`
        WHERE obs_date >= (SELECT date_sub(max(obs_date), INTERVAL 12 WEEK)
                           FROM `{marts}.stg_fred__h8`)
        ORDER BY obs_date, series_title
        """).to_dataframe()

    z_list = ", ".join(f"o.{c}" for c in Z_COLUMNS)
    movers = client.query_and_wait(f"""
        SELECT
            o.peer_band,
            o.cert,
            b.bank_name,
            round(o.composite_score, 3)                          AS composite,
            round(p.composite_score, 3)                          AS composite_prior,
            round(o.composite_score - p.composite_score, 3)      AS delta,
            {z_list}
        FROM `{marts}.mart_outlier_flags` o
        JOIN `{marts}.dim_banks` b USING (cert)
        JOIN `{marts}.mart_outlier_flags` p
          ON p.cert = o.cert AND p.report_date = '{prior}'
        WHERE o.report_date = '{latest}'
          AND o.composite_score IS NOT NULL AND p.composite_score IS NOT NULL
        """).to_dataframe()

    return {"latest": latest, "prior": prior, "kpis": kpis, "h8": h8, "movers": movers}


def _style_header(ws, row: int, width: int) -> None:
    for col in range(1, width + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def _autosize(ws, widths: dict[int, int]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width


def _pivot_data_sheet(wb: Workbook, movers: pd.DataFrame) -> int:
    """Every mover row as an Excel Table named MoversData — the source for
    the pivot sheet and for the workbook's live formulas. The delta column
    is a native formula (composite − prior), not a baked value."""
    ws = wb.create_sheet("PivotData")
    for j, col in enumerate(PIVOT_COLS, start=1):
        ws.cell(row=1, column=j, value=col)
    _style_header(ws, 1, len(PIVOT_COLS))
    delta_idx = PIVOT_COLS.index("delta") + 1
    for i, r in enumerate(movers[PIVOT_COLS].itertuples(index=False), start=2):
        for j, v in enumerate(r, start=1):
            if j == delta_idx:
                ws.cell(row=i, column=j, value=f"=D{i}-E{i}").number_format = "0.000"
            else:
                ws.cell(row=i, column=j,
                        value=None if pd.isna(v) else (v.item() if hasattr(v, "item") else v))
    n_rows = len(movers) + 1
    ref = f"A1:{get_column_letter(len(PIVOT_COLS))}{n_rows}"
    table = Table(displayName="MoversData", ref=ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(table)
    _autosize(ws, {1: 12, 2: 40, **{c: 12 for c in range(3, len(PIVOT_COLS) + 1)}})
    return n_rows


def _pivot_sheet(wb: Workbook, n_rows: int) -> None:
    """A real PivotTable over MoversData. No cache records are written — the
    cache is marked invalid + refresh-on-load, so Excel rebuilds it from the
    worksheet source the moment the file opens (verified against openpyxl
    3.1.5's writer; docs call creation unsupported, hence the belt-and-braces
    validation in the tests and the one-time open-in-Excel check)."""
    ncols = len(PIVOT_COLS)
    cache = CacheDefinition(
        invalid=True,
        refreshOnLoad=True,
        createdVersion=8, refreshedVersion=8, minRefreshableVersion=3,
        upgradeOnRefresh=True,
        recordCount=0,
        cacheSource=CacheSource(
            type="worksheet",
            worksheetSource=WorksheetSource(
                ref=f"A1:{get_column_letter(ncols)}{n_rows}", sheet="PivotData"),
        ),
        cacheFields=[
            CacheField(name="peer_band",
                       sharedItems=SharedItems(_fields=[Text(v=b) for b in BANDS])),
        ] + [CacheField(name=c, sharedItems=SharedItems()) for c in PIVOT_COLS[1:]],
    )
    band_items = [FieldItem(x=i) for i in range(len(BANDS))] + [FieldItem(t="default")]
    composite_idx = PIVOT_COLS.index("composite")
    delta_idx = PIVOT_COLS.index("delta")
    bank_idx = PIVOT_COLS.index("bank_name")
    pivot_fields = []
    for i in range(ncols):
        if i == 0:
            pivot_fields.append(PivotField(axis="axisRow", showAll=False, items=band_items))
        elif i in (composite_idx, delta_idx, bank_idx):
            pivot_fields.append(PivotField(dataField=True, showAll=False))
        else:
            pivot_fields.append(PivotField(showAll=False))
    pivot = TableDefinition(
        name="BandPivot",
        cacheId=1,
        dataCaption="Values",
        applyNumberFormats=False, applyBorderFormats=False, applyFontFormats=False,
        applyPatternFormats=False, applyAlignmentFormats=False,
        applyWidthHeightFormats=True,
        useAutoFormatting=True,
        itemPrintTitles=True,
        createdVersion=8, updatedVersion=8, minRefreshableVersion=3,
        indent=0, outline=True, outlineData=True, multipleFieldFilters=False,
        location=Location(ref="A3:D8", firstHeaderRow=1, firstDataRow=1, firstDataCol=1),
        pivotFields=pivot_fields,
        rowFields=[RowColField(x=0)],
        dataFields=[
            DataField(name="Average composite", fld=composite_idx,
                      subtotal="average", baseField=0, baseItem=0),
            DataField(name="Average change", fld=delta_idx,
                      subtotal="average", baseField=0, baseItem=0),
            DataField(name="Banks", fld=bank_idx,
                      subtotal="count", baseField=0, baseItem=0),
        ],
        pivotTableStyleInfo=PivotTableStyle(
            name="PivotStyleMedium9", showRowHeaders=True, showColHeaders=True,
            showRowStripes=False, showColStripes=False, showLastColumn=True),
    )
    pivot.cache = cache
    ws = wb.create_sheet("Pivot")
    ws["A1"] = "Composite score by size band — a live pivot over the PivotData sheet"
    ws["A1"].font = Font(bold=True, size=12)
    ws.add_pivot(pivot)
    _autosize(ws, {1: 24, 2: 18, 3: 18, 4: 12})


def build_workbook(frames: dict) -> Workbook:
    wb = Workbook()
    wb.calculation.fullCalcOnLoad = True

    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = f"Bank peer report — quarter ending {frames['latest']}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = ("Peer-relative statistics from public filings, never an "
                "assessment of any bank's condition.")
    kpis = frames["kpis"].iloc[0]
    labels = [
        ("Banks reporting", "banks_reporting"),
        ("Active banks", "active_banks"),
        ("Combined assets ($T)", "combined_assets_tn"),
        ("Median ROA (%)", "median_roa_pct"),
        ("Median NIM (%)", "median_nim_pct"),
        ("Median equity/assets (%)", "median_equity_pct"),
    ]
    for i, (label, key) in enumerate(labels, start=4):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws.cell(row=i, column=2, value=float(kpis[key]))

    ws.cell(row=11, column=1, value="Federal Reserve H.8, last 12 weeks ($B)").font = Font(bold=True)
    pivot = frames["h8"].pivot(index="obs_date", columns="series_title", values="value_billions")
    ws.cell(row=12, column=1, value="Week")
    for j, series in enumerate(pivot.columns, start=2):
        ws.cell(row=12, column=j, value=series)
    _style_header(ws, 12, len(pivot.columns) + 1)
    for i, (week, row) in enumerate(pivot.iterrows(), start=13):
        ws.cell(row=i, column=1, value=str(week))
        for j, value in enumerate(row, start=2):
            ws.cell(row=i, column=j, value=None if pd.isna(value) else float(value))

    # live formulas over the MoversData table — these cells compute in Excel,
    # nothing here is baked (XLOOKUP/MAXIFS need Excel's _xlfn prefix on disk;
    # Excel displays them unprefixed)
    r0 = 13 + len(pivot) + 2
    ws.cell(row=r0, column=1,
            value="By size band — live formulas over the PivotData table").font = Font(bold=True)
    for j, h in enumerate(["Band", "Banks", "Avg composite", "Top riser Δ", "Top riser"], start=1):
        ws.cell(row=r0 + 1, column=j, value=h)
    _style_header(ws, r0 + 1, 5)
    for i, band in enumerate(BANDS):
        r = r0 + 2 + i
        ws.cell(row=r, column=1, value=band)
        ws.cell(row=r, column=2, value=f'=COUNTIFS(MoversData[peer_band],A{r})')
        ws.cell(row=r, column=3,
                value=f'=ROUND(AVERAGEIFS(MoversData[composite],MoversData[peer_band],A{r}),3)')
        ws.cell(row=r, column=4,
                value=f'=ROUND(_xlfn.MAXIFS(MoversData[delta],MoversData[peer_band],A{r}),3)')
        ws.cell(row=r, column=5,
                value=(f'=_xlfn.XLOOKUP(_xlfn.MAXIFS(MoversData[delta],MoversData[peer_band],A{r}),'
                       f'MoversData[delta],MoversData[bank_name])'))
    _autosize(ws, {1: 28, 2: 22, 3: 22, 4: 22, 5: 40})

    for band in BANDS:
        ws = wb.create_sheet(band.replace(">", "over "))
        sub = frames["movers"][frames["movers"]["peer_band"] == band]
        risers = sub.sort_values("delta", ascending=False).head(MOVERS_PER_SIDE)
        fallers = sub.sort_values("delta").head(MOVERS_PER_SIDE)

        ws["A1"] = f"{band}: largest composite-score moves, {frames['prior']} → {frames['latest']}"
        ws["A1"].font = Font(bold=True, size=12)
        headers = ["Bank", "FDIC cert", "Composite", "Prior", "Change"] + list(Z_COLUMNS.values())
        row = 3
        for title, chunk in (("Risers", risers), ("Fallers", fallers)):
            ws.cell(row=row, column=1, value=title).font = Font(bold=True)
            row += 1
            for j, h in enumerate(headers, start=1):
                ws.cell(row=row, column=j, value=h)
            _style_header(ws, row, len(headers))
            first_data = row + 1
            for r in chunk.itertuples():
                row += 1
                # Change is a native formula over the two cells beside it
                values = [r.bank_name, int(r.cert), float(r.composite),
                          float(r.composite_prior), f"=C{row}-D{row}"]
                values += [None if pd.isna(getattr(r, c)) else round(float(getattr(r, c)), 2)
                           for c in Z_COLUMNS]
                for j, v in enumerate(values, start=1):
                    ws.cell(row=row, column=j, value=v)
                ws.cell(row=row, column=5).number_format = "0.000"
            if row >= first_data:
                z_range = (f"F{first_data}:{get_column_letter(5 + len(Z_COLUMNS))}{row}")
                ws.conditional_formatting.add(z_range, ColorScaleRule(
                    start_type="num", start_value=-5, start_color="4575B4",
                    mid_type="num", mid_value=0, mid_color="FFFFFF",
                    end_type="num", end_value=5, end_color="D73027",
                ))
            row += 2
        _autosize(ws, {1: 44, 2: 10, **{c: 12 for c in range(3, 6 + len(Z_COLUMNS))}})

    n_rows = _pivot_data_sheet(wb, frames["movers"])
    _pivot_sheet(wb, n_rows)

    ws = wb.create_sheet("Methodology")
    for i, line in enumerate(METHODOLOGY.strip().split("\n"), start=1):
        ws.cell(row=i, column=1, value=line)
    ws["A1"].font = Font(bold=True, size=12)
    ws.column_dimensions["A"].width = 78

    return wb


def quarter_tag(latest) -> str:
    d = pd.Timestamp(latest)
    return f"{d.year}Q{(d.month - 1) // 3 + 1}"


def main() -> int:
    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
    load_dotenv()
    project = os.environ.get("GCP_PROJECT")
    if not project:
        raise SystemExit("GCP_PROJECT is not set — see .env.example")
    marts = f"{project}.{os.environ.get('BQ_MARTS_DATASET', 'analytics')}"
    bucket_name = os.environ.get("GCS_REPORTS_BUCKET")

    client = bigquery.Client(project=project)
    try:
        frames = fetch_frames(client, marts)
    finally:
        client.close()
    wb = build_workbook(frames)
    name = f"bank_peer_report_{quarter_tag(frames['latest'])}.xlsx"

    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / name
        wb.save(path)
        log.info("built %s (%.1f KB)", name, path.stat().st_size / 1e3)
        if not bucket_name:
            out = Path(name)
            out.write_bytes(path.read_bytes())
            log.warning("GCS_REPORTS_BUCKET not set — wrote %s locally, skipped upload", out)
            return 0
        gcs = storage.Client(project=project)
        try:
            bucket = gcs.bucket(bucket_name)
            for blob_name in (f"reports/{name}", "reports/latest.xlsx"):
                bucket.blob(blob_name).upload_from_filename(str(path))
                log.info("uploaded gs://%s/%s", bucket_name, blob_name)
        finally:
            gcs.close()
    return 0


if __name__ == "__main__":
    sys.exit(main())
